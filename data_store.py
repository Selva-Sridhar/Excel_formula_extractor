import os
import json
import psycopg2
import psycopg2.extras as extras
from psycopg2 import pool
from io import StringIO
from typing import Dict, Any, List, Optional
from dotenv import load_dotenv
import time

load_dotenv()

# Connection pool (global)
connection_pool = None

def init_connection_pool():
    """Initialize PostgreSQL connection pool for efficient connections."""
    global connection_pool
    if connection_pool is None:
        connection_pool = pool.ThreadedConnectionPool(
            minconn=2,
            maxconn=10,
            host=os.getenv("PGHOST", "localhost"),
            port=int(os.getenv("PGPORT", "5432")),
            database=os.getenv("PGDATABASE", "excel_data_modified"),
            user=os.getenv("PGUSER", "postgres"),
            password=os.getenv("PGPASSWORD")
        )
        print("[INFO] Connection pool initialized")
    return connection_pool

def get_connection():
    """Get connection from pool with retry logic."""
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if connection_pool is None:
                init_connection_pool()
            conn = connection_pool.getconn()
            return conn
        except psycopg2.OperationalError as e:
            if attempt < max_retries - 1:
                print(f"[WARNING] Connection attempt {attempt + 1} failed, retrying...")
                time.sleep(2 ** attempt)
            else:
                print(f"[ERROR] Failed to connect after {max_retries} attempts: {e}")
                raise
    return None

def return_connection(conn):
    """Return connection to pool."""
    if connection_pool:
        connection_pool.putconn(conn)

def create_schema(conn):
    """Create optimized database schema."""
    cursor = conn.cursor()
    try:
        cursor.execute("""
        -- Table metadata
        CREATE TABLE IF NOT EXISTS table_metadata (
            id SERIAL PRIMARY KEY,
            file_name VARCHAR(500),
            sheet_name VARCHAR(255),
            table_name VARCHAR(255),
            table_type VARCHAR(50),
            range VARCHAR(100),
            row_count INTEGER DEFAULT 0,
            headers JSONB,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Table data (actual Excel values)
        CREATE TABLE IF NOT EXISTS table_data (
            id SERIAL PRIMARY KEY,
            metadata_id INTEGER REFERENCES table_metadata(id) ON DELETE CASCADE,
            row_number INTEGER,
            data JSONB,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        -- Formulas
        CREATE TABLE IF NOT EXISTS excel_formulas (
            id SERIAL PRIMARY KEY,
            file_name VARCHAR(500),
            sheet_name VARCHAR(255),
            cell_address VARCHAR(50),
            formula TEXT,
            readable_formula TEXT,
            dependencies JSONB,
            context JSONB,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """)
        conn.commit()
    except Exception as error:
        conn.rollback()
        print(f"[ERROR] Schema creation failed: {error}")
        raise
    finally:
        cursor.close()

def create_indexes(conn):
    """Create indexes after bulk data load for performance."""
    cursor = conn.cursor()
    try:
        cursor.execute("""
        CREATE INDEX IF NOT EXISTS idx_file_sheet ON table_metadata(file_name, sheet_name);
        CREATE INDEX IF NOT EXISTS idx_table_name ON table_metadata(table_name);
        CREATE INDEX IF NOT EXISTS idx_metadata_id ON table_data(metadata_id);
        CREATE INDEX IF NOT EXISTS idx_data_gin ON table_data USING GIN(data);
        CREATE INDEX IF NOT EXISTS idx_formula_file ON excel_formulas(file_name, sheet_name);
        CREATE INDEX IF NOT EXISTS idx_cell_address ON excel_formulas(cell_address);
        """)
        conn.commit()
    except Exception as error:
        conn.rollback()
        print(f"[ERROR] Index creation failed: {error}")
        raise
    finally:
        cursor.close()

def load_json_file(json_path: str) -> dict:
    """
    Load JSON file and return parsed data.

    Args:
        json_path: Path to JSON file

    Returns:
        dict: Parsed JSON data
    """
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"[INFO] Loaded JSON file: {json_path}")
        return data
    except Exception as e:
        print(f"[ERROR] Failed to load JSON file {json_path}: {e}")
        raise

def extract_and_store_tables_from_json(conn, tables_json_path: str, file_name: str, excel_file: str = None):
    """
    Extract tables from JSON and store in PostgreSQL.

    Args:
        conn: PostgreSQL connection
        tables_json_path: Path to tables JSON file
        file_name: Base filename for tracking
        excel_file: Optional path to Excel file for extracting actual data values

    Returns:
        dict: Mapping of (sheet_name, table_name) -> table_info for formula processing
    """
    import openpyxl

    # Load tables JSON
    tables_data = load_json_file(tables_json_path)

    cursor = conn.cursor()
    table_info_map = {}

    # Load Excel workbook if provided (for extracting actual data)
    wb = None
    if excel_file and os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        print(f"[INFO] Loaded Excel file for data extraction: {excel_file}")

    # Process each sheet
    for sheet_name, sheet_data in tables_data.items():
        print(f"[INFO] Processing sheet: {sheet_name}")

        ws = None
        if wb:
            ws = wb[sheet_name]

        # Process explicit tables
        for table in sheet_data.get("explicit_tables", []):
            table_name = table.get("table_name") or table.get("name")
            headers = table.get("headers", [])
            r1, c1, r2, c2 = table["r1"], table["c1"], table["r2"], table["c2"]
            range_str = table.get("range", "")

            # Extract data rows if Excel file is provided
            data_rows = []
            if ws:
                data_rows = extract_data_from_excel(ws, r1, r2, c1, c2, headers)

            # Insert metadata
            cursor.execute("""
            INSERT INTO table_metadata
            (file_name, sheet_name, table_name, table_type, range, row_count, headers)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """, (file_name, sheet_name, table_name, 'explicit',
                  range_str, len(data_rows), json.dumps(headers)))

            metadata_id = cursor.fetchone()[0]

            # Bulk insert data using COPY if data exists
            if data_rows:
                insert_table_data_copy_fast(cursor, metadata_id, data_rows)

            # Store for formula processing
            table_info_map[(sheet_name, table_name)] = {
                'r1': r1, 'c1': c1, 'r2': r2, 'c2': c2, 'headers': headers
            }

            print(f"  [✓] Explicit table '{table_name}': {len(data_rows)} rows")

        # Process implicit tables
        for table in sheet_data.get("implicit_tables", []):
            table_name = table.get("table_name")
            headers = table.get("header", [])
            r1, c1, r2, c2 = table["r1"], table["c1"], table["r2"], table["c2"]
            range_str = table.get("range", "")

            # Extract data rows if Excel file is provided
            data_rows = []
            if ws:
                data_rows = extract_data_from_excel(ws, r1, r2, c1, c2, headers)

            # Insert metadata
            cursor.execute("""
            INSERT INTO table_metadata
            (file_name, sheet_name, table_name, table_type, range, row_count, headers)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """, (file_name, sheet_name, table_name, 'implicit',
                  range_str, len(data_rows), json.dumps(headers)))

            metadata_id = cursor.fetchone()[0]

            # Bulk insert data
            if data_rows:
                insert_table_data_copy_fast(cursor, metadata_id, data_rows)

            # Store for formula processing
            table_info_map[(sheet_name, table_name)] = {
                'r1': r1, 'c1': c1, 'r2': r2, 'c2': c2, 'headers': headers
            }

            print(f"  [✓] Implicit table '{table_name}': {len(data_rows)} rows")

    if wb:
        wb.close()

    conn.commit()
    return table_info_map

def extract_data_from_excel(ws, r1, r2, c1, c2, headers):
    """
    Extract actual data from Excel worksheet.

    Args:
        ws: openpyxl worksheet
        r1, r2, c1, c2: Table boundaries
        headers: Column headers

    Returns:
        list: List of dictionaries containing row data
    """
    from openpyxl.utils import get_column_letter

    def get_merged_cell_value(ws, row, col):
        """Get value considering merged cells."""
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return ws.cell(row=min_row, column=min_col).value
        return ws.cell(row=row, column=col).value

    data_rows = []

    # Start from r1+1 to skip header row
    for row_idx in range(r1 + 1, r2 + 1):
        # Skip hidden rows
        if row_idx in ws.row_dimensions and ws.row_dimensions[row_idx].hidden:
            continue

        row_dict = {}
        is_empty = True

        for col_offset, header in enumerate(headers):
            col_idx = c1 + col_offset
            value = get_merged_cell_value(ws, row_idx, col_idx)

            if value is not None:
                is_empty = False
                # Handle datetime objects
                if hasattr(value, 'isoformat'):
                    value = value.isoformat()
                # Convert non-serializable types to string
                elif not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)

            row_dict[header] = value

        if not is_empty:
            data_rows.append(row_dict)

    return data_rows

def insert_table_data_copy_fast(cursor, metadata_id: int, data_rows: list):
    """Ultra-fast COPY insertion."""
    buffer = StringIO()
    for idx, row_data in enumerate(data_rows, start=1):
        json_data = json.dumps(row_data).replace('\\', '\\\\').replace('\n', '\\n')
        buffer.write(f"{metadata_id}|{idx}|{json_data}\n")

    buffer.seek(0)
    cursor.copy_from(buffer, 'table_data', sep='|',
                     columns=('metadata_id', 'row_number', 'data'))

def extract_and_store_formulas_from_json(conn, formulas_json_path: str, file_name: str):
    """
    Extract formulas from JSON and store in PostgreSQL.

    Args:
        conn: PostgreSQL connection
        formulas_json_path: Path to formulas JSON file
        file_name: Base filename
    """
    # Load formulas JSON
    formulas_data = load_json_file(formulas_json_path)

    cursor = conn.cursor()
    buffer = StringIO()
    formula_count = 0

    # Process formulas - the JSON contains a list of formula objects
    for formula_obj in formulas_data:
        cell_addr = formula_obj.get("cell", "")
        formula = formula_obj.get("formula", "")
        readable_formula = formula_obj.get("readable_formula", "")
        dependencies = formula_obj.get("dependencies", [])
        context = formula_obj.get("context", {})

        sheet_name = context.get("sheet", "")

        # Prepare for COPY
        def escape_csv(val):
            if val is None:
                return '\\N'
            return str(val).replace('\\', '\\\\').replace('\n', '\\n').replace('|', '\\|')

        buffer.write('|'.join([
            escape_csv(file_name),
            escape_csv(sheet_name),
            escape_csv(cell_addr),
            escape_csv(formula),
            escape_csv(readable_formula),
            json.dumps(dependencies),
            json.dumps(context)
        ]) + '\n')

        formula_count += 1

    # Bulk insert formulas using COPY
    if formula_count > 0:
        buffer.seek(0)
        cursor.copy_from(
            buffer,
            'excel_formulas',
            sep='|',
            columns=('file_name', 'sheet_name', 'cell_address', 'formula',
                     'readable_formula', 'dependencies', 'context')
        )
        print(f"  [✓] Inserted {formula_count} formulas")

    conn.commit()

def process_json_to_postgres(tables_json_path: str, formulas_json_path: str,
                              file_name: str, excel_file: str = None):
    """
    COMPLETE JSON TO POSTGRESQL PIPELINE.

    Args:
        tables_json_path: Path to tables JSON file
        formulas_json_path: Path to formulas JSON file
        file_name: Base filename for tracking
        excel_file: Optional path to Excel file for extracting actual data
    """
    # Get connection
    conn = get_connection()

    try:
        # Create schema
        create_schema(conn)

        print(f"\n[INFO] Processing tables from: {tables_json_path}")
        table_info_map = extract_and_store_tables_from_json(
            conn, tables_json_path, file_name, excel_file
        )

        print(f"\n[INFO] Processing formulas from: {formulas_json_path}")
        extract_and_store_formulas_from_json(conn, formulas_json_path, file_name)

        print(f"\n[INFO] Creating indexes...")
        create_indexes(conn)

        print(f"\n[SUCCESS] Data successfully stored in PostgreSQL!")

    except Exception as error:
        print(f"\n[ERROR] Pipeline failed: {error}")
        conn.rollback()
        raise
    finally:
        return_connection(conn)
