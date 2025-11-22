"""
Excel Table and Formula Extraction Tool
=======================================

This module provides functionality to:
1. Automatically detect and extract both explicit and implicit tables from Excel workbooks.
2. Extract and annotate Excel formulas with contextual header names.
3. Support both .xls and .xlsx files (with automatic conversion for .xls).

Key Features:
-------------
- Flood-fill–based implicit table detection using numpy (similar to the 'Number of Islands' logic).
- Header detection using text/numeric pattern analysis.
- Formula parsing and dependency mapping using the `formulas` library.
- Fallback to `xlwings` for formula extraction if `formulas` fails.
- JSON-based output for both table metadata and formula mapping.

"""


import os
import xlwings as xw
import formulas
import json
import datetime
import re
import openpyxl
import numpy as np
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import range_boundaries, get_column_letter


def ensure_xlsx(excel_path):
    """
    Ensures the Excel file is in .xlsx format.
    Converts .xls → .xlsx automatically using xlwings.

    Args:
        excel_path (str): Path to the Excel file.

    Returns:
        str: Path to the .xlsx file.
    """
    base, ext = os.path.splitext(excel_path)
    if ext.lower() == ".xls":
        print(f"[INFO] Detected .xls file: {excel_path}")
        converted_path = base + "_converted.xlsx"
        with xw.App(visible=False) as app:
            wb = app.books.open(excel_path)
            wb.save(converted_path)
            wb.close()
        print(f"[INFO] Converted .xls to .xlsx: {converted_path}")
        return converted_path
    return excel_path



def get_explicit_table_regions(ws):
    """
    Retrieves all explicit (Excel-defined) table regions in a worksheet.

    Args:
        ws (openpyxl.worksheet): The worksheet object.

    Returns:
        list: A list of tuples containing table boundaries and metadata:
              (min_row, min_col, max_row, max_col, table_name, table_object)
    """
    regions = []
    for t in ws._tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(t.ref)
        regions.append((min_row, min_col, max_row, max_col, t.displayName, t))
    return regions


def build_grid_excluding_explicit(ws, explicit_regions):
    """
    Builds a binary grid marking non-empty cells, excluding explicit Excel tables.

    Args:
        ws (openpyxl.worksheet): The worksheet.
        explicit_regions (list): List of explicit table regions to exclude.

    Returns:
        np.ndarray: A 2D boolean NumPy array where True marks non-empty cells.
    """
    rows = list(ws.iter_rows(values_only=True))
    R = len(rows)
    C = ws.max_column
    grid = np.zeros((R, C), dtype=bool)
    explicit_mask = np.zeros_like(grid, dtype=bool)
    for (r1, c1, r2, c2, _, _) in explicit_regions:
        explicit_mask[r1 - 1:r2, c1 - 1:c2] = True
    for i, row in enumerate(rows):
        for j in range(C):
            col_letter = get_column_letter(j + 1)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and getattr(col_dim, "hidden", False):
                continue
            val = row[j] if j < len(row) else None
            if val is not None and not explicit_mask[i, j]:
                grid[i, j] = True
    return grid


def flood_fill_islands(grid, min_rows=2, min_cols=2):
    """
    Identifies contiguous regions of non-empty cells (potential tables).

    Args:
        grid (np.ndarray): 2D boolean array representing non-empty cells.
        min_rows (int, optional): Minimum number of rows for a valid region.
        min_cols (int, optional): Minimum number of columns for a valid region.

    Returns:
        list: List of tuples representing table bounding boxes (r1, r2, c1, c2).
    """
    R, C = grid.shape
    visited = np.zeros_like(grid, dtype=bool)
    islands = []

    def flood(r, c):
        stack = [(r, c)]
        min_r = max_r = r
        min_c = max_c = c
        while stack:
            i, j = stack.pop()
            if i < 0 or i >= R or j < 0 or j >= C:
                continue
            if visited[i, j] or not grid[i, j]:
                continue
            visited[i, j] = True
            min_r = min(min_r, i)
            max_r = max(max_r, i)
            min_c = min(min_c, j)
            max_c = max(max_c, j)
            stack.extend([(i + 1, j), (i - 1, j), (i, j + 1), (i, j - 1)])
        return (min_r, max_r, min_c, max_c)

    for i in range(R):
        for j in range(C):
            if grid[i, j] and not visited[i, j]:
                min_r, max_r, min_c, max_c = flood(i, j)
                height = max_r - min_r + 1
                width = max_c - min_c + 1
                if height >= min_rows and width >= min_cols:
                    islands.append((min_r + 1, max_r + 1, min_c + 1, max_c + 1))
    return islands


def split_bbox_on_empty_lines(grid, bbox):
    """
    Splits a bounding box into smaller subregions if separated by empty rows or columns.

    Args:
        grid (np.ndarray): Boolean grid of non-empty cells.
        bbox (tuple): Bounding box (r1, r2, c1, c2).

    Returns:
        list: List of refined subregions as tuples.
    """
    r1, r2, c1, c2 = bbox
    r0, r1i = r1 - 1, r2 - 1
    c0, c1i = c1 - 1, c2 - 1
    subgrid = grid[r0:r1i + 1, c0:c1i + 1]
    row_sums = subgrid.sum(axis=1)
    col_sums = subgrid.sum(axis=0)

    empty_row_indices = [i for i, s in enumerate(row_sums) if s == 0]
    empty_col_indices = [j for j, s in enumerate(col_sums) if s == 0]

    if empty_col_indices:
        non_empty_cols = []
        start = None
        for j in range(subgrid.shape[1]):
            if col_sums[j] > 0:
                if start is None:
                    start = j
            else:
                if start is not None:
                    non_empty_cols.append((start, j - 1))
                    start = None
        if start is not None:
            non_empty_cols.append((start, subgrid.shape[1] - 1))
        subregions = []
        for (cs, ce) in non_empty_cols:
            new_c1 = c0 + cs + 1
            new_c2 = c0 + ce + 1
            subbbox = (r1, r2, new_c1, new_c2)
            subregions.extend(split_bbox_on_empty_rows(grid, subbbox))
        return subregions

    if empty_row_indices:
        non_empty_rows = []
        start = None
        for i in range(subgrid.shape[0]):
            if row_sums[i] > 0:
                if start is None:
                    start = i
            else:
                if start is not None:
                    non_empty_rows.append((start, i - 1))
                    start = None
        if start is not None:
            non_empty_rows.append((start, subgrid.shape[0] - 1))
        subregions = []
        for (rs, re) in non_empty_rows:
            new_r1 = r0 + rs + 1
            new_r2 = r0 + re + 1
            subregions.append((new_r1, new_r2, c1, c2))
        return subregions

    return [bbox]


def split_bbox_on_empty_rows(grid, bbox):
    """
    Helper function that recursively splits bounding boxes by empty rows/columns.

    Args:
        grid (np.ndarray): Boolean grid of non-empty cells.
        bbox (tuple): Bounding box (r1, r2, c1, c2).

    Returns:
        list: List of split bounding boxes.
    """
    r1, r2, c1, c2 = bbox
    r0, r1i = r1 - 1, r2 - 1
    c0, c1i = c1 - 1, c2 - 1
    subgrid = grid[r0:r1i + 1, c0:c1i + 1]
    row_sums = subgrid.sum(axis=1)
    col_sums = subgrid.sum(axis=0)

    empty_row_indices = [i for i, s in enumerate(row_sums) if s == 0]
    empty_col_indices = [j for j, s in enumerate(col_sums) if s == 0]

    if empty_row_indices:
        non_empty_rows = []
        start = None
        for i in range(subgrid.shape[0]):
            if row_sums[i] > 0:
                if start is None:
                    start = i
            else:
                if start is not None:
                    non_empty_rows.append((start, i - 1))
                    start = None
        if start is not None:
            non_empty_rows.append((start, subgrid.shape[0] - 1))
        out = []
        for rs, re in non_empty_rows:
            new_r1 = r0 + rs + 1
            new_r2 = r0 + re + 1
            out.extend(split_bbox_on_empty_lines(grid, (new_r1, new_r2, c1, c2)))
        return out

    if empty_col_indices:
        non_empty_cols = []
        start = None
        for j in range(subgrid.shape[1]):
            if col_sums[j] > 0:
                if start is None:
                    start = j
            else:
                if start is not None:
                    non_empty_cols.append((start, j - 1))
                    start = None
        if start is not None:
            non_empty_cols.append((start, subgrid.shape[1] - 1))
        out = []
        for cs, ce in non_empty_cols:
            new_c1 = c0 + cs + 1
            new_c2 = c0 + ce + 1
            out.append((r1, r2, new_c1, new_c2))
        return out

    return [bbox]


def get_merged_cell_value(ws, row, col):
    """
    Retrieves the value of a cell, considering merged cells.

    Args:
        ws (openpyxl.worksheet): Worksheet object.
        row (int): Row index.
        col (int): Column index.

    Returns:
        Any: Value from the cell or its merged range top-left cell.
    """
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return ws.cell(row=min_row, column=min_col).value
    return ws.cell(row=row, column=col).value


def detect_header_and_body(ws, r1, r2, c1, c2):
    """
    Attempts to detect the header row and data rows in a detected table region.

    Args:
        ws (openpyxl.worksheet): Worksheet object.
        r1, r2, c1, c2 (int): Bounding box coordinates.

    Returns:
        tuple: (header_row, body_rows)
    """
    rows = []
    for i in range(r1, r2 + 1):
        row = []
        for j in range(c1, c2 + 1):
            v = get_merged_cell_value(ws, i, j)
            row.append("" if v is None else str(v).strip())
        rows.append(row)

    def looks_text(s):
        return any(ch.isalpha() for ch in s)

    for idx in range(len(rows) - 1):
        text_count = sum(1 for cell in rows[idx] if cell and looks_text(cell))
        next_row = rows[idx + 1]
        numericish_next = sum(1 for cell in next_row if cell and not looks_text(cell))
        if text_count >= (len(rows[idx]) / 2) and numericish_next >= (len(next_row) / 3):
            return rows[idx], rows[idx + 1:]
    return rows[0], rows[1:]


def bbox_to_range_str(r1, r2, c1, c2):
    """
    Converts bounding box coordinates into Excel-style range string.

    Args:
        r1, r2, c1, c2 (int): Table coordinates.

    Returns:
        str: Excel range string (e.g., "A1:C10").
    """
    return f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"


def sanitize_table_headers_from_tableobj(ws, table_obj, r1, c1, r2, c2):
    """
    Cleans and retrieves valid headers from an explicit table or its first row.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet.
        table_obj (openpyxl.worksheet.table.Table): Table object.
        r1, c1, r2, c2 (int): Table boundaries.

    Returns:
        list: Cleaned list of header names.
    """
    headers = []
    try:
        cols = getattr(table_obj, "tableColumns", None)
        if cols:
            for col in cols:
                name = getattr(col, "name", "")
                if not name:
                    continue
                s = str(name).strip()
                if s == "" or s.startswith("[") or (not any(ch.isalpha() for ch in s)):
                    continue
                headers.append(s)
            if headers:
                return headers
    except Exception:
        pass

    header_row = []
    for j in range(c1, c2 + 1):
        col_letter = get_column_letter(j)
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim and getattr(col_dim, "hidden", False):
            continue
        v = get_merged_cell_value(ws, r1, j)
        s = "" if v is None else str(v).strip()
        if s == "" or s.startswith("[") or (not any(ch.isalpha() for ch in s)):
            continue
        header_row.append(s)
    return header_row


def generate_table_report(excel_file, out_json="Detected_Tables_report.json"):
    """
    Detects all explicit and implicit tables across all worksheets in an Excel file.

    Args:
        excel_file (str): Path to the Excel file.
        out_json (str, optional): Output JSON filename for detected tables.

    Returns:
        str: Path to the generated JSON file.
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    report = {}

    for ws in wb.worksheets:
        sheet_report = {"explicit_tables": [], "implicit_tables": []}
        table_counter = 1  # sequential names for tables without explicit names
        explicit = get_explicit_table_regions(ws)
        if explicit:
            for (r1, c1, r2, c2, name, table_obj) in explicit:
                headers = sanitize_table_headers_from_tableobj(ws, table_obj, r1, c1, r2, c2)
                effective_name = name if name else f"Table {table_counter}"
                sheet_report["explicit_tables"].append({
                    "name": name,
                    "table_name": effective_name,
                    "range": bbox_to_range_str(r1, r2, c1, c2),
                    "headers": headers,
                    "r1": r1, "c1": c1, "r2": r2, "c2": c2
                })
                table_counter += 1

        grid = build_grid_excluding_explicit(ws, explicit)
        islands = flood_fill_islands(grid)
        implicit_boxes = []
        for box in islands:
            parts = split_bbox_on_empty_lines(grid, box)
            for p in parts:
                implicit_boxes.append(p)

        if implicit_boxes:
            for (r1, r2, c1, c2) in implicit_boxes:
                header, _ = detect_header_and_body(ws, r1, r2, c1, c2)
                filtered_header = [h for h in header if h and not str(h).startswith("[")]
                sheet_report["implicit_tables"].append({
                    "table_name": f"Table {table_counter}",
                    "range": bbox_to_range_str(r1, r2, c1, c2),
                    "r1": r1, "c1": c1, "r2": r2, "c2": c2,
                    "header": filtered_header
                })
                table_counter += 1

        report[ws.title] = sheet_report

    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)
    print(f"Table extraction report saved to {out_json}")
    return out_json


# -------------------------------------------------
# Formula + Header Annotation
# -------------------------------------------------
def extract_references(formula):
    """
    Extracts all cell references from a given Excel formula.

    Args:
        formula (str): Excel formula string.

    Returns:
        list: List of cell references (e.g., ['A1', 'B2']).
    """
    pattern = r'(\$?[A-Za-z]{1,3}\$?\d+)'
    return re.findall(pattern, formula)


def find_header_for_cell(sheet_name, cell_ref, table_info):
    """
    Finds the header corresponding to a cell based on detected tables.

    Args:
        sheet_name (str): Sheet name.
        cell_ref (str): Cell reference (e.g., 'B2').
        table_info (dict): Parsed table metadata JSON.

    Returns:
        str or None: Header name if found, otherwise None.
    """
    try:
        col, row = coordinate_from_string(cell_ref.replace('$', ''))
        col_idx = column_index_from_string(col)
        row_idx = int(row)
    except Exception:
        return None

    tables = table_info.get(sheet_name, {})
    explicit_tables = tables.get("explicit_tables", [])
    for t in explicit_tables:
        r1 = t.get("r1")
        c1 = t.get("c1")
        r2 = t.get("r2")
        c2 = t.get("c2")
        if r1 <= row_idx <= r2 and c1 <= col_idx <= c2:
            header_pos = col_idx - c1
            headers = t.get("headers", [])
            if 0 <= header_pos < len(headers):
                return headers[header_pos]
    implicit_tables = tables.get("implicit_tables", [])
    for t in implicit_tables:
        r1 = t.get("r1")
        c1 = t.get("c1")
        r2 = t.get("r2")
        c2 = t.get("c2")
        if r1 <= row_idx <= r2 and c1 <= col_idx <= c2:
            header_pos = col_idx - c1
            headers = t.get("header", [])
            if 0 <= header_pos < len(headers):
                return headers[header_pos]
    return None


def annotate_formula(formula, sheet_name, table_info):
    """
    Annotates an Excel formula by replacing cell references with their header names.

    Args:
        formula (str): Original Excel formula.
        sheet_name (str): Sheet name.
        table_info (dict): Table info JSON data.

    Returns:
        str: Annotated formula string.
    """
    refs = set(extract_references(formula))
    annotated = formula
    refs_sorted = sorted(refs, key=len, reverse=True)
    for ref in refs_sorted:
        header = find_header_for_cell(sheet_name, ref, table_info)
        if header:
            annotated = re.sub(r'(?<![A-Za-z0-9_])' + re.escape(ref) + r'(?![A-Za-z0-9_])', f"[{header}]", annotated)
    return annotated


def extract_context(sheet, cell):
    """
    Extracts contextual metadata for a given Excel cell.

    Args:
        sheet (xlwings.Sheet): Worksheet object.
        cell (xlwings.Range): Excel cell.

    Returns:
        dict: Contextual information (sheet name, address, value).
    """
    return {
        "sheet": sheet.name,
        "cell_address": cell.address.replace('$', ''),
        "value": cell.value if cell.value is not None else "empty"
    }


class SafeEncoder(json.JSONEncoder):
    """
    Custom JSON encoder for datetime objects.
    """
    def default(self, obj):
        if isinstance(obj, (datetime.datetime, datetime.date)):
            return obj.isoformat()
        return super().default(obj)


def extract_formulas(excel_filepath, table_json_file, output_json_file):
    """
    Extracts formulas and their dependencies from an Excel workbook.

    Args:
        excel_filepath (str): Path to Excel workbook.
        table_json_file (str): Path to JSON file with detected table metadata.
        output_json_file (str): Path to save formula extraction output.

    Returns:
        None: Writes output to JSON file.

    Notes:
        - Uses formulas library when available.
        - Falls back to xlwings if formulas library fails.
        - Each formula is annotated with header context.
    """
    excel_filepath = ensure_xlsx(excel_filepath)

    with open(table_json_file, "r", encoding="utf-8") as f:
        table_info = json.load(f)

    formula_records = []

    try:
        excel_model = formulas.ExcelModel().loads(excel_filepath).finish()
        formula_mode = "formulas"
    except Exception:
        print("WARNING: Formulas library failed, falling back to xlwings-only mode.")
        excel_model = None
        formula_mode = "xlwings"

    with xw.App(visible=False) as app:
        wb = app.books.open(excel_filepath)
        for sheet in wb.sheets:
            sheet_name = sheet.name
            for cell in sheet.used_range:
                if cell.formula and isinstance(cell.formula, str) and cell.formula.startswith('='):
                    formula = cell.formula
                    context = extract_context(sheet, cell)
                    readable_formula = annotate_formula(formula, sheet_name, table_info)

                    dependencies = []
                    if excel_model:
                        addr = f"'[{wb.name}]{sheet_name.upper()}'!{cell.address.replace('$', '')}"
                        cell_obj = excel_model.cells.get(addr)
                        if cell_obj and getattr(cell_obj, "inputs", None):
                            deps = list(cell_obj.inputs.keys())
                            simple_deps = []
                            for d in deps:
                                m = re.search(r'([A-Za-z]{1,3}\$?\d+:[A-Za-z]{1,3}\$?\d+)', d)
                                simple_deps.append(m.group(1) if m else d)
                            dependencies = list(dict.fromkeys(simple_deps))
                        else:
                            dependencies = list(dict.fromkeys(extract_references(formula)))
                    else:
                        dependencies = list(dict.fromkeys(extract_references(formula)))

                    formula_records.append({
                        "cell": cell.address.replace('$', ''),
                        "formula": formula,
                        "readable_formula": readable_formula,
                        "dependencies": dependencies,
                        "context": context
                    })
        wb.close()

    with open(output_json_file, "w", encoding="utf-8") as f:
        json.dump(formula_records, f, indent=2, cls=SafeEncoder)
    print(f"Extracted formulas with header context saved to {output_json_file}")

import os

def run_analysis(excel_filepath,
                 table_json_path="outputs/Detected_Tables_report.json",
                 formula_output_path="outputs/excel_formulas_structured.json"):

    # ✅ Extract clean file name
    base = os.path.basename(excel_filepath)              # "3_statement_excel_completed_model.xlsx"
    name_only = os.path.splitext(base)[0]                # "3_statement_excel_completed_model"

    # ✅ Append clean name to output file paths
    table_json_path = f"outputs/{name_only}_tables.json"
    formula_output_path = f"outputs/{name_only}_formulas.json"

    # ✅ Continue with the workflow
    excel_filepath = ensure_xlsx(excel_filepath)
    generate_table_report(excel_filepath, table_json_path)
    extract_formulas(excel_filepath, table_json_path, formula_output_path)
